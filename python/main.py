from elasticsearch import Elasticsearch
from elasticsearch import helpers
import pandas as pd
import json as json
import csv
import xlsxwriter
import openpyxl
import numpy as np


from datetime import datetime, date, time, timedelta
import calendar

"""autor: Santiago Quiroz Upegui (squiroz@unal.edu.co)
"""
#obtiene los datos en un intervalo de tiempo y los guarda en un arreglo 
def getInterval(es,inferior, superior):
    
    cres = helpers.scan(es,index="test-index", query={"query": {
    "bool": {
      "must": [
        {
          "match_all": {}
        },
        {
          "range": {
            "dt": {
              "gte": inferior.strftime('%Y-%m-%d %H:%M:%S'),
              "lte": superior.strftime('%Y-%m-%d %H:%M:%S'),
              "format": "yyyy-MM-dd HH:mm:ss||epoch_millis"
            }
          }
        }
      ],
      "filter": [],
      "should": [],
      "must_not": []
    }}},preserve_order=True)
    crestresults=[]
    for hit in cres:
        crestresults.append(hit)
    
    print(len(crestresults))
    return crestresults
#obtiene arreglos de intervalos de tiempo, cuenta cuantos datos hay en cada intervalo y los guarda en disco segun la ruta dada
def getTotal(es,inferior,fechaFinal,intervaloSegundos,directorio):
    intervalo=[]
    cantidad=[]
    datos=[]
    superior=datetime.now()
    while inferior < fechaFinal:
        #print("inferior: "+inferior.strftime('%Y-%m-%d %H:%M:%S'))
        superior =inferior + timedelta(seconds=intervaloSegundos)
        if superior > fechaFinal :
            superior = fechaFinal
        #print("superior: "+superior.strftime('%Y-%m-%d %H:%M:%S'))
        result=getInterval(es,inferior,superior)
        intervalo.append((inferior.strftime('%Y-%m-%d %H:%M:%S')+'//'+superior.strftime('%Y-%m-%d %H:%M:%S')))
        cantidad.append(len(result))
        datos.append([result])
        inferior=superior
    
    #escribiendo en disco
    
    entradas=len(cantidad)
    #df = pd.DataFrame({'Intervalo': intervalo,'VEIzquierda':[None]*entradas,'VEDerecha':[None]*entradas,'Etiquetado':[None]*entradas,'Cantidad': cantidad,'Datos':datos})
    df = pd.DataFrame({'Intervalo': intervalo,'Etiquetado':[None]*entradas,'Cantidad': cantidad,'Datos':datos})
    print(df)
    print("Guardando en disco.")
    writer = pd.ExcelWriter(directorio+".xlsx", engine='xlsxwriter')
    df.to_excel(writer,index_label = 'indice', sheet_name='hoja1')
    writer.save()
    df = pd.DataFrame({'Intervalo': intervalo,'Etiquetado':[None]*entradas,'Cantidad': cantidad})
    writer = pd.ExcelWriter(directorio+"SINDATOS.xlsx", engine='xlsxwriter')
    df.to_excel(writer,index_label = 'indice', sheet_name='hoja1')
    writer.save()
    return True
def getTrainingData(es,intervaloSegundos = 5):
    directorio = "./Entrenamiento/training"
    fechaInicial= datetime.strptime('2019-03-20 12:00:00', '%Y-%m-%d %H:%M:%S')
    fechaFinal= datetime.strptime('2019-03-24 12:00:00', '%Y-%m-%d %H:%M:%S')
    print("inicial: "+fechaInicial.strftime('%Y-%m-%d %H:%M:%S'))
    print("final: "+fechaFinal.strftime('%Y-%m-%d %H:%M:%S'))
    getTotal(es,fechaInicial,fechaFinal,intervaloSegundos,directorio)
    return True
    
def getActualData(es,intervaloSegundos = 5):
    directorio = "./Evaluacion/evaluation"
    fechaFinal= datetime.now()
    fechaInicial= fechaFinal-timedelta(hours=4)
    print("inicial: "+fechaInicial.strftime('%Y-%m-%d %H:%M:%S'))
    print("final: "+fechaFinal.strftime('%Y-%m-%d %H:%M:%S'))
    getTotal(es,fechaInicial,fechaFinal,intervaloSegundos,directorio)
    
    return True
def clasificar(url,numeroAnteriores=30):
  doc = openpyxl.load_workbook(url)
  hoja=doc.get_sheet_by_name("hoja1")
  #.cell(row=1,column=1).value
  cambioAnteriores=[]
  anterior = hoja.cell(row=2,column=4).value
  hoja.cell(row=2,column=3).value = "normal"  
  siguiente = hoja.cell(row=3,column=4).value
  hoja.cell(row=3,column=3).value = "normal"
  cambio=0
  if siguiente != None:
    cambio=abs(int(anterior)-int(siguiente))/int(anterior)
    cambioAnteriores.append(cambio)
  else:
    print("Error: campo nulo.")
    return  True
  print(anterior)
  print(cambio)
  print(cambioAnteriores)
  print(np.average(cambioAnteriores))
  i=4
  anterior = siguiente
  siguiente =siguiente= hoja.cell(row=i,column=4).value
  while(siguiente != None):
    
    cambio=abs(int(anterior)-int(siguiente))/int(anterior)
    if(cambio>(np.average(cambioAnteriores)+0.2)):
      hoja.cell(row=i,column=3).value = "anormal"
      cambio=cambioAnteriores[-1]
      if(len(cambioAnteriores)==numeroAnteriores):
        cambioAnteriores=cambioAnteriores[1:]
        cambioAnteriores.append(cambio)
      else:
        cambioAnteriores.append(cambio) 
    else:
      hoja.cell(row=i,column=3).value = "normal" 
      if(len(cambioAnteriores)==numeroAnteriores):
        cambioAnteriores=cambioAnteriores[1:]
        cambioAnteriores.append(cambio)
      else:
        cambioAnteriores.append(cambio)

    anterior = siguiente
    i = i+1
    siguiente= hoja.cell(row=i,column=4).value
  nombre=url.split('.')
  doc.save("."+nombre[1]+"-ETIQUETADOS.xlsx")
  return True
def etiquetar():
    cual=int(input("Que datos desea etiquetar? \n 1.entrenamiento. \n 2.evaluacion. \n 3.todos los anteriores.\n"))
    if (cual == 1):
      clasificar("./Entrenamiento/trainingSINDATOS.xlsx")
    elif (cual == 2):
      clasificar("./Entrenamiento/trainingSINDATOS.xlsx")
    elif (cual == 3):
      clasificar("./Entrenamiento/trainingSINDATOS.xlsx")
      clasificar("./Evaluacion/evaluationSINDATOS.xlsx")
    else:
      print("Ha ingresado una opcion no valida")
    return True
if __name__ == "__main__":
    es = Elasticsearch(['54.147.121.162'],
    #http_auth=('user', 'secret'), 
    scheme="http",
    port=9200,
    timeout=30)
    while(True):
        opcion= int(input("Elija una opcion: \n 1.Obtener datos de entrenamiento. \n 2.obtener datos de evaluacion. \n 3.obtener datos de entrenamiento y evaluacion. \n 4.Etiquetar.\n 5.Salir.\n"))
        if opcion == 1:
            getTrainingData(es)
        elif ( opcion == 2):
            getActualData(es)
        elif ( opcion ==3):
            getTrainingData(es)
            getActualData(es)
        elif ( opcion == 4):
            etiquetar()
        elif ( opcion == 5):
            break
        else:
          print("Ha ingresado una opcion no valida")        
